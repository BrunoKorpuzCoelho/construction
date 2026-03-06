# -*- coding: utf-8 -*-
import base64
import io
import logging
import re
from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  numbers as xl_numbers)
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    _logger.warning("openpyxl não instalado — instalar com: pip install openpyxl")


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORT WIZARD
# ═══════════════════════════════════════════════════════════════════════════════
class BOQExportWizard(models.TransientModel):
    _name = 'construction.boq.export.wizard'
    _description = 'Exportar BOQ para Excel'

    boq_id = fields.Many2one('construction.boq', 'BOQ', required=True)
    export_type = fields.Selection([
        ('internal', 'Exportação Interna (completa)'),
        ('client', 'Exportação para Cliente (simplificada)'),
    ], 'Tipo de Exportação', required=True, default='client')
    include_obs = fields.Boolean('Incluir Observações', default=True)
    include_stock = fields.Boolean('Incluir Coluna Stock')
    file_data = fields.Binary('Ficheiro', readonly=True)
    file_name = fields.Char('Nome do Ficheiro', readonly=True)
    state = fields.Selection([('draft', 'Configurar'), ('done', 'Exportado')], default='draft')

    def action_export(self):
        if not HAS_OPENPYXL:
            raise UserError(_('Por favor instale openpyxl: pip install openpyxl'))
        self.ensure_one()
        wb = self._build_workbook()
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        obra_ref = self.boq_id.obra_id.ref_interna or 'BOQ'
        fname = f'{obra_ref}_{self.boq_id.revision_label}_{"cliente" if self.export_type == "client" else "interno"}.xlsx'
        self.write({
            'file_data': base64.b64encode(buf.read()),
            'file_name': fname,
            'state': 'done',
        })
        return {'type': 'ir.actions.act_window', 'res_model': self._name,
                'res_id': self.id, 'view_mode': 'form', 'target': 'new'}

    def _build_workbook(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'BOQ'

        # ── Styles ──────────────────────────────────────────────────────────
        style_title  = {'font': Font(bold=True, size=14, color='FFFFFF'),
                         'fill': PatternFill('solid', fgColor='714B67'),
                         'alignment': Alignment(horizontal='left', vertical='center')}
        style_cap    = {'font': Font(bold=True, size=11, color='FFFFFF'),
                         'fill': PatternFill('solid', fgColor='1E3A5F'),
                         'alignment': Alignment(horizontal='left', vertical='center')}
        style_sub    = {'font': Font(bold=True, size=10, color='BAD8F5'),
                         'fill': PatternFill('solid', fgColor='243B55'),
                         'alignment': Alignment(horizontal='left', vertical='center')}
        style_header = {'font': Font(bold=True, size=10, color='CBD5E1'),
                         'fill': PatternFill('solid', fgColor='334155'),
                         'alignment': Alignment(horizontal='center', vertical='center')}
        style_total  = {'font': Font(bold=True, size=11, color='FFFFFF'),
                         'fill': PatternFill('solid', fgColor='0F2340'),
                         'alignment': Alignment(horizontal='right', vertical='center')}

        thin = Side(style='thin', color='E2E8F0')
        border = Border(bottom=thin)

        def apply_style(cell, style):
            for k, v in style.items():
                setattr(cell, k, v)

        # ── Columns ──────────────────────────────────────────────────────────
        cols = ['Código', 'Descrição do Artigo', 'Un.', 'Quantidade', 'P.U. (€)', 'Total (€)']
        if self.include_obs:
            cols.append('Observações')
        if self.export_type == 'internal' and self.include_stock:
            cols.append('Stock')
        col_count = len(cols)

        # ── Title row ────────────────────────────────────────────────────────
        obra = self.boq_id.obra_id
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        t = ws.cell(1, 1, f'{obra.display_name} — {self.boq_id.name} {self.boq_id.revision_label}')
        apply_style(t, style_title)
        ws.row_dimensions[1].height = 26

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
        ws.cell(2, 1, f'Exportado em: {fields.Datetime.now().strftime("%d/%m/%Y %H:%M")} | Estado: {self.boq_id.state}')
        ws.cell(2, 1).font = Font(italic=True, size=9, color='94A3B8')

        # ── Header row ───────────────────────────────────────────────────────
        hrow = 4
        for ci, h in enumerate(cols, 1):
            c = ws.cell(hrow, ci, h)
            apply_style(c, style_header)
        ws.row_dimensions[hrow].height = 20

        # ── Load data via SQL ─────────────────────────────────────────────────
        cr = self.env.cr
        cr.execute("""
            SELECT
                s.id AS sec_id, s.parent_id, s.code AS sec_code,
                s.name AS sec_name, s.depth, s.path, s.is_leaf,
                s.specialty,
                a.code, a.name, a.qty_contract, a.price_unit,
                a.qty_contract * a.price_unit AS total, a.obs,
                u.name AS uom_name,
                COALESCE(sq.qty, 0) AS stock_qty
            FROM construction_boq_section s
            LEFT JOIN construction_boq_artigo a
                   ON a.section_id = s.id AND a.active = TRUE
            LEFT JOIN uom_uom u ON u.id = a.uom_id
            LEFT JOIN (
                SELECT sq2.product_id, SUM(sq2.quantity) AS qty
                FROM stock_quant sq2
                JOIN stock_location sl ON sl.id = sq2.location_id AND sl.usage = 'internal'
                GROUP BY sq2.product_id
            ) sq ON sq.product_id = a.product_id
            WHERE s.boq_id = %s
            ORDER BY s.path, a.sequence
        """, (self.boq_id.id,))
        rows = cr.dictfetchall()

        # ── Write data ───────────────────────────────────────────────────────
        row = hrow + 1
        sec_totals = {}   # sec_id -> running total
        grand_total = 0.0
        seen_sections = set()

        for r in rows:
            sec_id = r['sec_id']

            # Write section header on first encounter
            if sec_id not in seen_sections:
                seen_sections.add(sec_id)
                indent = '  ' * r['depth']
                label = f"{indent}{r['sec_code']}  {r['sec_name'].upper() if r['depth'] == 0 else r['sec_name']}"
                sec_style = style_cap if r['depth'] == 0 else style_sub
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
                c = ws.cell(row, 1, label)
                apply_style(c, sec_style)
                ws.row_dimensions[row].height = 18 if r['depth'] == 0 else 16
                row += 1
                sec_totals[sec_id] = 0.0

            # Write article row if article data present
            if r['name']:
                total = float(r['qty_contract'] or 0) * float(r['price_unit'] or 0)
                ws.cell(row, 1, r['code'] or '').border = border
                ws.cell(row, 2, r['name'] or '').border = border
                ws.cell(row, 3, r['uom_name'] or '').border = border
                ws.cell(row, 3).alignment = Alignment(horizontal='center')

                qty_cell = ws.cell(row, 4, float(r['qty_contract'] or 0))
                qty_cell.number_format = '#,##0.000'
                qty_cell.alignment = Alignment(horizontal='right')
                qty_cell.border = border

                pu_cell = ws.cell(row, 5, float(r['price_unit'] or 0))
                pu_cell.number_format = '#,##0.0000 €'
                pu_cell.alignment = Alignment(horizontal='right')
                pu_cell.border = border

                tot_cell = ws.cell(row, 6, total)
                tot_cell.number_format = '#,##0.00 €'
                tot_cell.alignment = Alignment(horizontal='right')
                tot_cell.font = Font(bold=True)
                tot_cell.border = border

                ci = 7
                if self.include_obs:
                    ws.cell(row, ci, r['obs'] or '').border = border; ci += 1
                if self.export_type == 'internal' and self.include_stock:
                    ws.cell(row, ci, float(r['stock_qty'] or 0)).border = border

                sec_totals[sec_id] = sec_totals.get(sec_id, 0.0) + total
                grand_total += total
                row += 1

        # Grand Total
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count - 1)
        c = ws.cell(row, 1, 'TOTAL GERAL (s/ IVA)')
        apply_style(c, style_total)
        c.font = Font(bold=True, size=12, color='FFFFFF')
        gt = ws.cell(row, col_count, grand_total)
        apply_style(gt, style_total)
        gt.number_format = '#,##0.00 €'
        gt.font = Font(bold=True, size=12, color='FFFFFF')
        ws.row_dimensions[row].height = 22

        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count - 1)
        ws.cell(row, 1, 'TOTAL c/ IVA 23%').font = Font(italic=True, color='64748B')
        iva_cell = ws.cell(row, col_count, grand_total * 1.23)
        iva_cell.number_format = '#,##0.00 €'
        iva_cell.font = Font(italic=True, color='64748B')
        iva_cell.alignment = Alignment(horizontal='right')

        # ── Column widths ────────────────────────────────────────────────────
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 55
        ws.column_dimensions['C'].width = 7
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 16
        if self.include_obs:
            ws.column_dimensions[get_column_letter(7)].width = 25

        ws.freeze_panes = f'A{hrow + 1}'

        return wb


# ═══════════════════════════════════════════════════════════════════════════════
# IMPORT WIZARD
# ═══════════════════════════════════════════════════════════════════════════════
class BOQImportWizard(models.TransientModel):
    _name = 'construction.boq.import.wizard'
    _description = 'Importar BOQ de Excel'

    boq_id = fields.Many2one('construction.boq', 'BOQ Destino', required=True)
    file_data = fields.Binary('Ficheiro Excel', required=True)
    file_name = fields.Char('Nome do Ficheiro')
    mode = fields.Selection([
        ('replace', 'Substituir tudo'),
        ('append', 'Adicionar ao existente'),
    ], 'Modo', required=True, default='replace')
    preview_lines = fields.Text('Pré-visualização', readonly=True)
    state = fields.Selection([('upload', 'Upload'), ('preview', 'Pré-visualizar'),
                               ('done', 'Concluído')], default='upload')
    import_log = fields.Text('Log de Importação', readonly=True)

    def action_preview(self):
        if not HAS_OPENPYXL:
            raise UserError(_('Por favor instale openpyxl: pip install openpyxl'))
        rows = self._parse_excel()
        preview = f"Encontradas {len(rows)} linhas de artigos.\n"
        paths = set()
        for r in rows:
            if r.get('section_path'):
                paths.add(r['section_path'][-1][0] if r['section_path'] else '')
        preview += f"Secções detectadas: {', '.join(sorted(paths))}\n"
        preview += "\nPrimeiras 5 linhas:\n"
        for r in rows[:5]:
            preview += f"  {r.get('code','')} | {r.get('name','')[:40]} | {r.get('uom','')} | {r.get('qty',0)} | {r.get('pu',0)}\n"
        self.write({'preview_lines': preview, 'state': 'preview'})
        return {'type': 'ir.actions.act_window', 'res_model': self._name,
                'res_id': self.id, 'view_mode': 'form', 'target': 'new'}

    def action_import(self):
        if not HAS_OPENPYXL:
            raise UserError(_('Por favor instale openpyxl: pip install openpyxl'))
        boq = self.boq_id
        if boq.state in ('shared', 'archived'):
            raise UserError(_('Não é possível importar para um BOQ bloqueado.'))

        rows = self._parse_excel()
        log = self._do_import(boq, rows)
        self.write({'state': 'done', 'import_log': log})
        return {'type': 'ir.actions.act_window', 'res_model': self._name,
                'res_id': self.id, 'view_mode': 'form', 'target': 'new'}

    def _parse_excel(self):
        """
        Reads Excel and returns list of dicts with article data.
        Each dict has:
          section_path: list of (code, name) tuples from root to leaf section
          code, name, uom, qty, pu, obs
        """
        data = base64.b64decode(self.file_data)
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active

        rows = []
        current_path = []  # list of (code, name) tuples

        HEADER_PATTERNS = re.compile(r'^(código|code|descrição|description|un\.|qty|p\.u\.|total)', re.I)

        for row in ws.iter_rows(values_only=True):
            if not any(row):
                continue
            vals = [str(c).strip() if c is not None else '' for c in row]

            if HEADER_PATTERNS.match(vals[0] if vals else ''):
                continue

            # Detect section header row (no qty, no price, has code)
            if vals[0] and not vals[3] and not vals[4]:
                code = vals[0].strip()
                name = vals[1].strip() if len(vals) > 1 else ''
                # Determine depth by number of dots in code
                depth = code.count('.')
                # Trim current_path to this depth
                current_path = current_path[:depth]
                current_path.append((code, name))
                continue

            # Try to parse as article row
            try:
                code = vals[0] if len(vals) > 0 else ''
                name = vals[1] if len(vals) > 1 else ''
                uom  = vals[2] if len(vals) > 2 else ''
                qty  = float(str(vals[3]).replace(',', '.') or 0) if len(vals) > 3 and vals[3] else 0
                pu   = float(str(vals[4]).replace(',', '.') or 0) if len(vals) > 4 and vals[4] else 0
                obs  = vals[6] if len(vals) > 6 else ''

                if not name:
                    continue

                rows.append({
                    'section_path': list(current_path),
                    'code': code,
                    'name': name,
                    'uom': uom,
                    'qty': qty,
                    'pu': pu,
                    'obs': obs,
                })
            except (ValueError, IndexError):
                continue

        wb.close()
        return rows

    def _do_import(self, boq, rows):
        """Insert BOQ data via parameterised SQL using the new section model."""
        cr = self.env.cr
        uid = self.env.uid

        if self.mode == 'replace':
            cr.execute("DELETE FROM construction_boq_artigo WHERE boq_id = %s", (boq.id,))
            cr.execute("DELETE FROM construction_boq_section WHERE boq_id = %s", (boq.id,))

        # section_cache: tuple of codes -> section_id
        section_cache = {}
        log_lines = []
        art_count = 0

        uom_cache = {}
        for uom in self.env['uom.uom'].search([]):
            uom_cache[uom.name.lower().strip()] = uom.id

        for r in rows:
            section_path = r.get('section_path', [])
            if not section_path:
                log_lines.append(f'SKIP: {r["code"]} — sem secção definida')
                continue

            # Walk section_path top-down, creating missing sections
            parent_id = None
            for idx, (code, name) in enumerate(section_path):
                path_key = tuple(c for c, _ in section_path[:idx + 1])
                if path_key in section_cache:
                    parent_id = section_cache[path_key]
                    continue

                # SELECT existing
                if parent_id:
                    cr.execute(
                        "SELECT id FROM construction_boq_section "
                        "WHERE boq_id=%s AND parent_id=%s AND code=%s LIMIT 1",
                        (boq.id, parent_id, code))
                else:
                    cr.execute(
                        "SELECT id FROM construction_boq_section "
                        "WHERE boq_id=%s AND parent_id IS NULL AND code=%s LIMIT 1",
                        (boq.id, code))
                existing = cr.fetchone()
                if existing:
                    sec_id = existing[0]
                else:
                    # Compute depth and path
                    depth = idx
                    if parent_id:
                        cr.execute("SELECT path FROM construction_boq_section WHERE id=%s",
                                   (parent_id,))
                        parent_path = cr.fetchone()[0] or ''
                    else:
                        parent_path = ''

                    cr.execute("""
                        SELECT COALESCE(MAX(sequence), 0) + 10
                        FROM construction_boq_section
                        WHERE boq_id=%s AND (parent_id=%s OR (parent_id IS NULL AND %s IS NULL))
                    """, (boq.id, parent_id, parent_id))
                    seq = cr.fetchone()[0]
                    seg = str(seq).zfill(4)
                    path = f"{parent_path}.{seg}" if parent_path else seg

                    is_leaf = (idx == len(section_path) - 1)
                    cr.execute("""
                        INSERT INTO construction_boq_section
                            (boq_id, parent_id, code, name, sequence, depth, path,
                             is_leaf, specialty, color, notes,
                             create_uid, write_uid, create_date, write_date)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'General','#1E3A5F','',
                                %s,%s,NOW(),NOW())
                        RETURNING id
                    """, (boq.id, parent_id, code, name or code, seq, depth, path,
                          is_leaf, uid, uid))
                    sec_id = cr.fetchone()[0]

                    # Mark parent as non-leaf
                    if parent_id:
                        cr.execute("""
                            UPDATE construction_boq_section
                            SET is_leaf=FALSE, write_uid=%s, write_date=NOW()
                            WHERE id=%s
                        """, (uid, parent_id))

                section_cache[path_key] = sec_id
                parent_id = sec_id

            leaf_sec_id = parent_id
            if not leaf_sec_id:
                log_lines.append(f'SKIP: {r["code"]} — secção folha não encontrada')
                continue

            uom_id = uom_cache.get((r.get('uom') or '').lower().strip())

            cr.execute("""
                INSERT INTO construction_boq_artigo
                    (boq_id, section_id, code, name,
                     uom_id, qty_contract, price_unit, obs, sequence, active,
                     create_uid, write_uid, create_date, write_date)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s,
                    (SELECT COALESCE(MAX(sequence),0)+10 FROM construction_boq_artigo WHERE section_id=%s),
                    TRUE, %s, %s, NOW(), NOW())
            """, (boq.id, leaf_sec_id,
                  r['code'], r['name'], uom_id,
                  float(r.get('qty', 0)), float(r.get('pu', 0)),
                  r.get('obs', ''), leaf_sec_id,
                  uid, uid))
            art_count += 1

        # Final is_leaf update: sections with children → is_leaf=False
        cr.execute("""
            UPDATE construction_boq_section s
            SET is_leaf = FALSE
            WHERE s.boq_id = %s
              AND EXISTS (
                SELECT 1 FROM construction_boq_section c WHERE c.parent_id = s.id
              )
        """, (boq.id,))

        log_lines.insert(0, f'Importados com sucesso: {art_count} artigos')
        return '\n'.join(log_lines)
