# -*- coding: utf-8 -*-
import base64
import io
import logging
import re
from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  numbers as xl_numbers)
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    _logger.warning("openpyxl não instalado — instalar com: pip install openpyxl")


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORT WIZARD
# ═══════════════════════════════════════════════════════════════════════════════
class BOQExportWizard(models.TransientModel):
    _name = 'construction.boq.export.wizard'
    _description = 'Exportar BOQ para Excel'

    boq_id = fields.Many2one('construction.boq', 'BOQ', required=True)
    export_type = fields.Selection([
        ('internal', 'Exportação Interna (completa)'),
        ('client', 'Exportação para Cliente (simplificada)'),
    ], 'Tipo de Exportação', required=True, default='client')
    include_obs = fields.Boolean('Incluir Observações', default=True)
    include_stock = fields.Boolean('Incluir Coluna Stock')
    file_data = fields.Binary('Ficheiro', readonly=True)
    file_name = fields.Char('Nome do Ficheiro', readonly=True)
    state = fields.Selection([('draft', 'Configurar'), ('done', 'Exportado')], default='draft')

    def action_export(self):
        if not HAS_OPENPYXL:
            raise UserError(_('Por favor instale openpyxl: pip install openpyxl'))
        self.ensure_one()
        wb = self._build_workbook()
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        obra_ref = self.boq_id.obra_id.ref_interna or 'BOQ'
        fname = f'{obra_ref}_{self.boq_id.revision_label}_{"cliente" if self.export_type == "client" else "interno"}.xlsx'
        self.write({
            'file_data': base64.b64encode(buf.read()),
            'file_name': fname,
            'state': 'done',
        })
        return {'type': 'ir.actions.act_window', 'res_model': self._name,
                'res_id': self.id, 'view_mode': 'form', 'target': 'new'}

    def _build_workbook(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'BOQ'

        # ── Styles ──────────────────────────────────────────────────────────
        style_title  = {'font': Font(bold=True, size=14, color='FFFFFF'),
                         'fill': PatternFill('solid', fgColor='714B67'),
                         'alignment': Alignment(horizontal='left', vertical='center')}
        style_cap    = {'font': Font(bold=True, size=11, color='FFFFFF'),
                         'fill': PatternFill('solid', fgColor='1E3A5F'),
                         'alignment': Alignment(horizontal='left', vertical='center')}
        style_sub    = {'font': Font(bold=True, size=10, color='BAD8F5'),
                         'fill': PatternFill('solid', fgColor='243B55'),
                         'alignment': Alignment(horizontal='left', vertical='center')}
        style_header = {'font': Font(bold=True, size=10, color='CBD5E1'),
                         'fill': PatternFill('solid', fgColor='334155'),
                         'alignment': Alignment(horizontal='center', vertical='center')}
        style_total  = {'font': Font(bold=True, size=11, color='FFFFFF'),
                         'fill': PatternFill('solid', fgColor='0F2340'),
                         'alignment': Alignment(horizontal='right', vertical='center')}
        style_money  = {'alignment': Alignment(horizontal='right')}
        style_qty    = {'alignment': Alignment(horizontal='right')}

        thin = Side(style='thin', color='E2E8F0')
        border = Border(bottom=thin)

        def apply_style(cell, style):
            for k, v in style.items():
                setattr(cell, k, v)

        # ── Columns ──────────────────────────────────────────────────────────
        cols = ['Código', 'Descrição do Artigo', 'Un.', 'Quantidade', 'P.U. (€)', 'Total (€)']
        if self.include_obs:
            cols.append('Observações')
        if self.export_type == 'internal' and self.include_stock:
            cols.append('Stock')
        col_count = len(cols)

        # ── Title row ────────────────────────────────────────────────────────
        obra = self.boq_id.obra_id
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        t = ws.cell(1, 1, f'{obra.display_name} — {self.boq_id.name} {self.boq_id.revision_label}')
        apply_style(t, style_title)
        ws.row_dimensions[1].height = 26

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
        ws.cell(2, 1, f'Exportado em: {fields.Datetime.now().strftime("%d/%m/%Y %H:%M")} | Estado: {self.boq_id.state}')
        ws.cell(2, 1).font = Font(italic=True, size=9, color='94A3B8')

        # ── Header row ───────────────────────────────────────────────────────
        hrow = 4
        for ci, h in enumerate(cols, 1):
            c = ws.cell(hrow, ci, h)
            apply_style(c, style_header)
        ws.row_dimensions[hrow].height = 20

        # ── Load data via SQL ─────────────────────────────────────────────────
        cr = self.env.cr
        cr.execute("""
            SELECT
                c.code AS cap_code, c.name AS cap_name, c.specialty,
                sc.code AS sub_code, sc.name AS sub_name,
                a.code, a.name, a.qty_contract, a.price_unit,
                a.qty_contract * a.price_unit AS total, a.obs,
                u.name AS uom_name,
                COALESCE(sq.qty, 0) AS stock_qty
            FROM construction_boq_capitulo c
            JOIN construction_boq_subcapitulo sc ON sc.capitulo_id = c.id
            JOIN construction_boq_artigo a ON a.subcapitulo_id = sc.id AND a.active = TRUE
            LEFT JOIN uom_uom u ON u.id = a.uom_id
            LEFT JOIN (
                SELECT sq2.product_id, SUM(sq2.quantity) AS qty
                FROM stock_quant sq2
                JOIN stock_location sl ON sl.id = sq2.location_id AND sl.usage = 'internal'
                GROUP BY sq2.product_id
            ) sq ON sq.product_id = a.product_id
            WHERE c.boq_id = %s
            ORDER BY c.sequence, sc.sequence, a.sequence
        """, (self.boq_id.id,))
        rows = cr.dictfetchall()

        # ── Write data ───────────────────────────────────────────────────────
        current_cap = None
        current_sub = None
        row = hrow + 1
        cap_total = 0
        sub_total = 0
        grand_total = 0

        for r in rows:
            # New capítulo
            if r['cap_code'] != current_cap:
                if current_sub is not None:
                    self._write_subtotal(ws, row, col_count, current_sub, sub_total, style_sub)
                    row += 1; sub_total = 0; current_sub = None
                if current_cap is not None:
                    self._write_captotal(ws, row, col_count, current_cap, cap_total, style_cap)
                    row += 1; cap_total = 0
                current_cap = r['cap_code']
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
                c = ws.cell(row, 1, f"{r['cap_code']}  {r['cap_name'].upper()}")
                apply_style(c, style_cap)
                ws.row_dimensions[row].height = 18
                row += 1

            # New subcapítulo
            if r['sub_code'] != current_sub:
                if current_sub is not None:
                    self._write_subtotal(ws, row, col_count, current_sub, sub_total, style_sub)
                    row += 1; sub_total = 0
                current_sub = r['sub_code']
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count - 1)
                c = ws.cell(row, 1, f"  {r['sub_code']}  {r['sub_name']}")
                apply_style(c, style_sub)
                ws.row_dimensions[row].height = 16
                row += 1

            # Artigo row
            total = float(r['qty_contract'] or 0) * float(r['price_unit'] or 0)
            ws.cell(row, 1, r['code'] or '').border = border
            ws.cell(row, 2, r['name'] or '').border = border
            ws.cell(row, 3, r['uom_name'] or '').border = border
            ws.cell(row, 3).alignment = Alignment(horizontal='center')

            qty_cell = ws.cell(row, 4, float(r['qty_contract'] or 0))
            qty_cell.number_format = '#,##0.000'
            qty_cell.alignment = Alignment(horizontal='right')
            qty_cell.border = border

            pu_cell = ws.cell(row, 5, float(r['price_unit'] or 0))
            pu_cell.number_format = '#,##0.0000 €'
            pu_cell.alignment = Alignment(horizontal='right')
            pu_cell.border = border

            tot_cell = ws.cell(row, 6, total)
            tot_cell.number_format = '#,##0.00 €'
            tot_cell.alignment = Alignment(horizontal='right')
            tot_cell.font = Font(bold=True)
            tot_cell.border = border

            ci = 7
            if self.include_obs:
                ws.cell(row, ci, r['obs'] or '').border = border; ci += 1
            if self.export_type == 'internal' and self.include_stock:
                ws.cell(row, ci, float(r['stock_qty'] or 0)).border = border; ci += 1

            sub_total += total; cap_total += total; grand_total += total
            row += 1

        # Flush last sub/cap
        if current_sub:
            self._write_subtotal(ws, row, col_count, current_sub, sub_total, style_sub); row += 1
        if current_cap:
            self._write_captotal(ws, row, col_count, current_cap, cap_total, style_cap); row += 1

        # Grand Total
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count - 1)
        c = ws.cell(row, 1, 'TOTAL GERAL (s/ IVA)')
        apply_style(c, style_total)
        c.font = Font(bold=True, size=12, color='FFFFFF')
        gt = ws.cell(row, col_count, grand_total)
        apply_style(gt, style_total)
        gt.number_format = '#,##0.00 €'
        gt.font = Font(bold=True, size=12, color='FFFFFF')
        ws.row_dimensions[row].height = 22

        # IVA row
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count - 1)
        ws.cell(row, 1, 'TOTAL c/ IVA 23%').font = Font(italic=True, color='64748B')
        iva_cell = ws.cell(row, col_count, grand_total * 1.23)
        iva_cell.number_format = '#,##0.00 €'
        iva_cell.font = Font(italic=True, color='64748B')
        iva_cell.alignment = Alignment(horizontal='right')

        # ── Column widths ────────────────────────────────────────────────────
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 55
        ws.column_dimensions['C'].width = 7
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 16
        if self.include_obs:
            ws.column_dimensions[get_column_letter(7)].width = 25

        # ── Freeze panes ─────────────────────────────────────────────────────
        ws.freeze_panes = f'A{hrow + 1}'

        return wb

    def _write_subtotal(self, ws, row, col_count, sub_code, total, style):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count - 1)
        c = ws.cell(row, 1, f'    Subtotal {sub_code}')
        for k, v in style.items():
            setattr(c, k, v)
        c.font = Font(bold=True, italic=True, size=10, color='BAD8F5')
        t = ws.cell(row, col_count, total)
        t.number_format = '#,##0.00 €'
        t.font = Font(bold=True, color='60A5FA')
        t.alignment = Alignment(horizontal='right')
        for k, v in style.items():
            if k != 'font': setattr(t, k, v)
        ws.row_dimensions[row].height = 15

    def _write_captotal(self, ws, row, col_count, cap_code, total, style):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count - 1)
        c = ws.cell(row, 1, f'TOTAL {cap_code}')
        for k, v in style.items():
            setattr(c, k, v)
        c.font = Font(bold=True, size=11, color='FFFFFF')
        t = ws.cell(row, col_count, total)
        t.number_format = '#,##0.00 €'
        t.font = Font(bold=True, size=11, color='FFFFFF')
        t.alignment = Alignment(horizontal='right')
        for k, v in style.items():
            if k != 'font': setattr(t, k, v)
        ws.row_dimensions[row].height = 18


# ═══════════════════════════════════════════════════════════════════════════════
# IMPORT WIZARD
# ═══════════════════════════════════════════════════════════════════════════════
class BOQImportWizard(models.TransientModel):
    _name = 'construction.boq.import.wizard'
    _description = 'Importar BOQ de Excel'

    boq_id = fields.Many2one('construction.boq', 'BOQ Destino', required=True)
    file_data = fields.Binary('Ficheiro Excel', required=True)
    file_name = fields.Char('Nome do Ficheiro')
    mode = fields.Selection([
        ('replace', 'Substituir tudo'),
        ('append', 'Adicionar ao existente'),
    ], 'Modo', required=True, default='replace')
    preview_lines = fields.Text('Pré-visualização', readonly=True)
    state = fields.Selection([('upload', 'Upload'), ('preview', 'Pré-visualizar'),
                               ('done', 'Concluído')], default='upload')
    import_log = fields.Text('Log de Importação', readonly=True)

    def action_preview(self):
        if not HAS_OPENPYXL:
            raise UserError(_('Por favor instale openpyxl: pip install openpyxl'))
        rows = self._parse_excel()
        preview = f"Encontradas {len(rows)} linhas de artigos.\n"
        caps = set(r['cap_code'] for r in rows if r.get('cap_code'))
        preview += f"Capítulos detectados: {', '.join(sorted(caps))}\n"
        preview += "\nPrimeiras 5 linhas:\n"
        for r in rows[:5]:
            preview += f"  {r.get('code','')} | {r.get('name','')[:40]} | {r.get('uom','')} | {r.get('qty',0)} | {r.get('pu',0)}\n"
        self.write({'preview_lines': preview, 'state': 'preview'})
        return {'type': 'ir.actions.act_window', 'res_model': self._name,
                'res_id': self.id, 'view_mode': 'form', 'target': 'new'}

    def action_import(self):
        if not HAS_OPENPYXL:
            raise UserError(_('Por favor instale openpyxl: pip install openpyxl'))
        boq = self.boq_id
        if boq.state in ('shared', 'archived'):
            raise UserError(_('Não é possível importar para um BOQ bloqueado.'))

        rows = self._parse_excel()
        log = self._do_import(boq, rows)
        self.write({'state': 'done', 'import_log': log})
        return {'type': 'ir.actions.act_window', 'res_model': self._name,
                'res_id': self.id, 'view_mode': 'form', 'target': 'new'}

    def _parse_excel(self):
        """Lê o Excel e devolve lista de dicts com os dados de cada artigo."""
        data = base64.b64decode(self.file_data)
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active

        rows = []
        current_cap_code = ''
        current_cap_name = ''
        current_sub_code = ''
        current_sub_name = ''

        HEADER_PATTERNS = re.compile(r'^(código|code|descrição|description|un\.|qty|p\.u\.|total)', re.I)

        for row in ws.iter_rows(values_only=True):
            if not any(row):
                continue
            vals = [str(c).strip() if c is not None else '' for c in row]

            # Detect header row — skip
            if HEADER_PATTERNS.match(vals[0] if vals else ''):
                continue

            # Detect capítulo rows (merged, no numeric in col 4/5)
            if vals[0] and not vals[3] and not vals[4]:
                # Could be cap or sub based on indentation / code pattern
                code = vals[0]
                name = vals[1] if len(vals) > 1 else ''
                if re.match(r'^\d{1,3}$', code.strip()):
                    current_cap_code = code.strip()
                    current_cap_name = name.strip()
                    current_sub_code = ''
                    current_sub_name = ''
                elif re.match(r'^\d{1,3}\.\d{2}$', code.strip()):
                    current_sub_code = code.strip()
                    current_sub_name = name.strip()
                continue

            # Try to parse as artigo row
            try:
                code = vals[0] if len(vals) > 0 else ''
                name = vals[1] if len(vals) > 1 else ''
                uom  = vals[2] if len(vals) > 2 else ''
                qty  = float(str(vals[3]).replace(',', '.') or 0) if len(vals) > 3 and vals[3] else 0
                pu   = float(str(vals[4]).replace(',', '.') or 0) if len(vals) > 4 and vals[4] else 0
                obs  = vals[6] if len(vals) > 6 else ''

                if not name:
                    continue

                # Infer cap/sub from code if not already set
                if code and re.match(r'^\d{1,3}\.\d{2}\.\d{1,4}', code):
                    parts = code.split('.')
                    if not current_cap_code:
                        current_cap_code = parts[0]
                    if not current_sub_code:
                        current_sub_code = f'{parts[0]}.{parts[1]}'

                rows.append({
                    'cap_code': current_cap_code,
                    'cap_name': current_cap_name,
                    'sub_code': current_sub_code,
                    'sub_name': current_sub_name,
                    'code': code,
                    'name': name,
                    'uom': uom,
                    'qty': qty,
                    'pu': pu,
                    'obs': obs,
                })
            except (ValueError, IndexError):
                continue

        wb.close()
        return rows

    def _do_import(self, boq, rows):
        """Insert BOQ data via parameterised SQL. No ON CONFLICT — SELECT first."""
        cr = self.env.cr
        uid = self.env.uid

        if self.mode == 'replace':
            # Hard-delete everything so the import starts from a clean slate.
            # Order matters: artigos → subcapitulos → capitulos (FK constraints).
            cr.execute("DELETE FROM construction_boq_artigo WHERE boq_id = %s", (boq.id,))
            cr.execute("DELETE FROM construction_boq_subcapitulo WHERE boq_id = %s", (boq.id,))
            cr.execute("DELETE FROM construction_boq_capitulo WHERE boq_id = %s", (boq.id,))

        cap_map = {}  # cap_code -> cap_id
        sub_map = {}  # "cap.sub" -> sub_id
        log_lines = []
        art_count = 0

        # Pre-cache all UoM ids — avoids 200k ORM calls for large imports
        uom_cache = {}
        for uom in self.env['uom.uom'].search([]):
            uom_cache[uom.name.lower().strip()] = uom.id

        for r in rows:
            cap_key = r['cap_code']
            sub_key = f"{r['cap_code']}.{r['sub_code']}" if r['sub_code'] else None

            # ── Chapter: SELECT first, INSERT if missing ──────────────────
            if cap_key and cap_key not in cap_map:
                cr.execute(
                    "SELECT id FROM construction_boq_capitulo WHERE boq_id=%s AND code=%s LIMIT 1",
                    (boq.id, cap_key))
                existing = cr.fetchone()
                if existing:
                    cap_map[cap_key] = existing[0]
                else:
                    cr.execute("""
                        INSERT INTO construction_boq_capitulo
                            (boq_id, code, name, sequence, specialty,
                             create_uid, write_uid, create_date, write_date)
                        VALUES (%s, %s, %s,
                            (SELECT COALESCE(MAX(sequence),0)+10
                             FROM construction_boq_capitulo WHERE boq_id=%s),
                            'General', %s, %s, NOW(), NOW())
                        RETURNING id
                    """, (boq.id, cap_key, r['cap_name'] or cap_key,
                          boq.id, uid, uid))
                    cap_map[cap_key] = cr.fetchone()[0]

            # ── Sub-chapter: SELECT first, INSERT if missing ───────────────
            if sub_key and sub_key not in sub_map and cap_key in cap_map:
                cap_id = cap_map[cap_key]
                cr.execute(
                    "SELECT id FROM construction_boq_subcapitulo "
                    "WHERE capitulo_id=%s AND code=%s LIMIT 1",
                    (cap_id, r['sub_code']))
                existing = cr.fetchone()
                if existing:
                    sub_map[sub_key] = existing[0]
                else:
                    cr.execute("""
                        INSERT INTO construction_boq_subcapitulo
                            (boq_id, capitulo_id, code, name, sequence,
                             create_uid, write_uid, create_date, write_date)
                        VALUES (%s, %s, %s, %s,
                            (SELECT COALESCE(MAX(sequence),0)+10
                             FROM construction_boq_subcapitulo WHERE capitulo_id=%s),
                            %s, %s, NOW(), NOW())
                        RETURNING id
                    """, (boq.id, cap_id, r['sub_code'], r['sub_name'] or r['sub_code'],
                          cap_id, uid, uid))
                    sub_map[sub_key] = cr.fetchone()[0]

            cap_id = cap_map.get(cap_key)
            sub_id = sub_map.get(sub_key)
            if not cap_id or not sub_id:
                log_lines.append(f'SKIP: {r["code"]} — cap/sub não determinados')
                continue

            # Lookup UoM from pre-cached dict
            uom_id = uom_cache.get((r.get('uom') or '').lower().strip())

            cr.execute("""
                INSERT INTO construction_boq_artigo
                    (boq_id, capitulo_id, subcapitulo_id, code, name,
                     uom_id, qty_contract, price_unit, obs, sequence, active,
                     create_uid, write_uid, create_date, write_date)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s,
                    (SELECT COALESCE(MAX(sequence),0)+10 FROM construction_boq_artigo WHERE subcapitulo_id=%s),
                    TRUE, %s, %s, NOW(), NOW())
            """, (boq.id, cap_id, sub_id,
                  r['code'], r['name'], uom_id,
                  float(r.get('qty', 0)), float(r.get('pu', 0)),
                  r.get('obs', ''), sub_id,
                  self.env.uid, self.env.uid))
            art_count += 1

        log_lines.insert(0, f'Importados com sucesso: {art_count} artigos')
        return '\n'.join(log_lines)
